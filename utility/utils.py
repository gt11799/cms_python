#!/usr/bin/env python
# -*- coding: utf-8 -*-

#!/usr/bin/python
# coding:utf-8

import os
import sys

# path = os.path.dirname(__file__)
# sys.path.append(os.path.join(path, ".."))
from os.path import dirname,abspath
system_root = dirname(dirname( abspath( __file__ ) ) ) #定义上层目录为根目录 
import sys;sys.path.insert(0,system_root)  #把项目根目录加入默认库路径 

from settings import *
try:
    import MySQLdb
except:
    pass
import redis
import pymongo
import traceback
import logging
import sys
from utility.error_code import *
import copy
import types
import datetime
import time
from msg_code import MSG_CODE
from mail import send_mail
from qiniu.models import getDefineSizeImage
import os
import inspect
import tornado.gen
import motor

# import MySQLdb
from MySQLdb.cursors import DictCursor
from DBUtils.PooledDB import PooledDB


import xlrd
from  openpyxl.reader.excel  import  load_workbook  

import random

logging.basicConfig(level=logging.DEBUG)

DB_TIME_FORMAT = '%Y-%m-%d %H:%M:%S'
DB_DATE_FORMAT = '%Y-%m-%d'
client = getClient()

class MyHumanDict(dict):
    
    def __init__(self,*arg,**kw):
        super(MyHumanDict, self).__init__(*arg, **kw)

    def __getitem__(self, key):
        self.get(key,None)

def LimitIPQueryCount(count=200):
    def temp(func):
        def deal(self,*args,**kwargs):
            client_ip = self.request.headers.get("Ali-Cdn-Real-Ip") or self.request.remote_ip
            request = self.__class__.__name__
            date = str(datetime.date.today())
            conn = getMongoDBConn()
            db = conn.shop
            obj = db.query_limit.find_and_modify({"date":date,"request":request,"client_ip":client_ip},
                    {"$inc":{"count":1}},True,new=True)
            if obj["count"] > count:
                # raise MyDefineError("尝试次数过多，请稍后再尝试.")
                self.write({"status":2,"msg":"尝试次数过多，请稍后再尝试."})
                return
            else:
                return func(self,*args,**kwargs)
        return deal
    return temp        


def errorLog(show_caller=False):
    try:
        from settings import ERROR_LOG_PATH
        ch = logging.FileHandler(filename=ERROR_LOG_PATH)
    except:
        if os.path.exists("/root/log/"):
            if client == 'web':
                ch = logging.FileHandler(filename='/root/log/error_web.log')
            else:
                ch = logging.FileHandler(filename='/root/log/error.log')
        else:
            ch = logging.StreamHandler()
    ch.setLevel(logging.ERROR)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    log = logging.getLogger('ERROR')
    log.propagate = False
    log.addHandler(ch)
    if show_caller:
        frm = inspect.stack()[1]
        finfo = inspect.getframeinfo(frm[0])
        caller_info =  """ 
        who call errordebug function:
        filename:%s
        function:%s
        lineno:%s
        """%(finfo.filename,finfo.function,finfo.lineno)
        log.error(caller_info)
    return log

def debugLog():
    if os.path.exists("/root/log/"):
        if client == "web":
            ch = logging.FileHandler(filename='/root/log/debug_web.log')
        else:
            ch = logging.FileHandler(filename='/root/log/debug.log')
    else:
        ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    log = logging.getLogger('DEBUG')
    log.propagate = False
    log.addHandler(ch)

    return log

ERRORLOG = errorLog()

DEBUGLOG = debugLog()

def countTime(func):
    def temp(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        logStr =  "%s spend %s" % (func.__name__, end - start)
        logging.error(logStr)
        return result
    return temp


# @countTime
def getRedisObj(rdb=0):

    # r = redis.Redis(password=REDIS_PASSWD,
                    # db=rdb, unix_socket_path='/tmp/redis.sock')
    key = "REDIS_POOL_%s"%rdb
    pool = globals().get(key)
    if not pool:
        pool = redis.ConnectionPool(
            host=REDIS_HOST, password=REDIS_PASSWD, port=6379, db=rdb)
        globals()[key] = pool

    # print id(pool)
    r = redis.Redis(connection_pool=pool)
    return r



def static_url_patch():
 
    from tornado.template  import Template
    old_generate = Template.generate
    def hack_generate(self,**kwargs):
        t = old_generate(self,**kwargs)
        import re
        c = re.compile("/static/")
        # print "hack generate"
        t = re.sub(r'(?<=[\'|"])/static/',r'http://cdn.xiaoher.com/static/',t)
        return t
 
    tornado.template.Template.generate = hack_generate

try:
    from settings import STATIC_CDN_ON
    if STATIC_CDN_ON:
        static_url_patch()
except:
    pass





def getUid(token):
    return -1


def SQLInsertBuilder(objDict, tableName):
    sql = "insert into " + tableName + "("
    for key in objDict.keys():
        sql += key + ","
    sql = sql.rstrip(",")
    sql += ") values("
    for key in objDict.keys():
        print objDict[key]
        if type(objDict[key]) == type(1):
            sql += "'" + str(objDict[key]) + "',"
        else:
            sql += "'" + objDict[key] + "',"

    sql = sql.rstrip(",")
    sql = sql + ");"
    return sql


def SQLReplaceBuilderEx(objDict, tableName):
    sql = "replace into " + tableName + "("
    for key in objDict.keys():
        sql += key + ","
    sql = sql.rstrip(",")
    sql += ") values("
    for key in objDict.keys():
        if type(objDict[key]) == type(1):
            sql += "'" + str(objDict[key]) + "',"
        else:
            sql += "'" + objDict[key] + "',"

    sql = sql.rstrip(",")
    sql = sql + ");"
    return sql

def NewSQLInsertBuilder(objDict,tableName,update=False):
    sql = "insert into " + tableName + "("
    for key in objDict.keys():
        sql += key + ","
    sql = sql.rstrip(",")
    sql += ") values("
    for key in objDict.keys():
        # print objDict[key]
        if isinstance(objDict[key],int) or isinstance(objDict[key],float):
            sql += "%s"%objDict[key]
        else:
            sql += "'%s'"%objDict[key]
            
        sql += ","

    sql = sql.rstrip(",")
    sql = sql + ")"
    if update:
        sql += " ON DUPLICATE KEY UPDATE "
        for key in objDict.keys():
            if isinstance(objDict[key],int) or isinstance(objDict[key],float):
                sql += "%s=%s,"%(key,objDict[key])
            else:
                sql += "%s='%s',"%(key,objDict[key])

        sql = sql.rstrip(",")
    
    return sql


def timestamp2Str(stamp):
    x = time.localtime(stamp)
    return time.strftime('%Y-%m-%d %H:%M:%S', x)


def time2Str(timeObj):
    return timeObj.strftime('%Y-%m-%d %H:%M:%S')


def SQLReplaceBuilder(objDict, tableName):
    sql = "replace into " + tableName + "("
    for key in objDict.keys():
        sql += key + ","
    sql = sql.rstrip(",")
    sql += ") values("
    for key in objDict.keys():
        if type(objDict[key]) == types.IntType:
            sql += "'" + str(objDict[key]) + "',"
        else:
            sql += "'" + objDict[key] + "',"

    sql = sql.rstrip(",")
    sql = sql + ");"
    return sql


class DBAccess:
    dbName = ''
    transDb = None
    multiDb = None
    multiCursor = None
    passwd = DB_PWD
    host = DB_HOST
    query_host = DB_HOST

    def __init__(self, app_type=''):
        reload(sys)
        sys.setdefaultencoding('utf-8')

    def _conn(self,dbName):
        conn = MySQLdb.connect(
            user='root', db=dbName, passwd=self.passwd, host=DB_HOST, charset='utf8')
        return conn


    def execNonQuery(self, sql, args=None):
        lastId = 0
        db = MySQLdb.connect(
            user='root', db=self.dbName, passwd=self.passwd, host=DB_HOST, charset='utf8')
        cursor = db.cursor()
        rowCount = cursor.execute(sql, args)
        lastId = int(db.insert_id())
        db.commit()
        cursor.close()
        db.close()
        return lastId

    def execUpdate(self, sql):
        db = MySQLdb.connect(
            user='root', db=self.dbName, passwd=DB_PWD, host=DB_HOST, charset='utf8')
        cursor = db.cursor()
        rowCount = cursor.execute(sql)
        db.commit()
        cursor.close()
        db.close()
        return rowCount

    def execQuery(self, sql ,slave = True):
        db = MySQLdb.connect(
        user='root', db=self.dbName, passwd=DB_PWD, host=DB_HOST, charset='utf8')
        cursor = db.cursor()
        cursor.execute(sql)
        db.commit()
        cds = cursor.fetchall()
        cursor.close()
        db.close()
        return cds

    def execQueryAssoc(self, sql ,slave = True):
        db = MySQLdb.connect(
        user='root', db=self.dbName, passwd=DB_PWD, host=DB_HOST, charset='utf8')
        cursor = db.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(sql)
        db.commit()
        cds = cursor.fetchall()
        cursor.close()
        db.close()
        return cds

    def multiInit(self):
        self.multiDb = MySQLdb.connect(
            user='root', db=self.dbName, passwd=DB_PWD, host=DB_HOST, charset='utf8')
        self.multiCursor = self.multiDb.cursor()

    def multiQuery(self, sql):
        self.multiCursor.execute(sql)
        self.multiDb.commit()
        cds = self.multiCursor.fetchall()
        return cds

    def multiNonQuery(self, sql):
        self.multiCursor.execute(sql)
        self.multiDb.commit()

    def multiEnd(self):
        self.multiCursor.close()
        self.multiDb.close()
    #

    def beginTrans(self):
        self.transDb = MySQLdb.connect(
            user='root', db=self.dbName, passwd=DB_PWD, host=DB_HOST, charset='utf8')
        self.transDb.autocommit(False)

    def transNonQuery(self, sql):
        cursor = self.transDb.cursor()
        rowCount = cursor.execute(sql)
        cursor.close()
        return rowCount

    def endTrans(self):
        self.transDb.commit()
        self.transDb.close()

    def rollback(self):
        self.transDb.rollback()
        self.transDb.close()

class DBPool(object):
    """数据库连接池"""
    #连接池对象
    poolDict = {}
    poolSelectDict = {} #读写分离
    @staticmethod
    def getConn(dbName):
        """
        @summary: 静态方法，从连接池中取出连接
        @return MySQLdb.connection
        """
        if not DBPool.poolDict.has_key(dbName):
            pool = PooledDB(creator=MySQLdb, mincached=1 , maxcached=1 ,
                              host=DB_HOST , port=DB_PORT , user=DB_USER , passwd=DB_PWD ,
                              db=dbName,use_unicode=False,charset='utf8',cursorclass=DictCursor)
            DBPool.poolDict[dbName] = pool
        return DBPool.poolDict[dbName].connection()

    @staticmethod
    def getSelectConn(dbName):
        """
        @summary: 静态方法，从连接池中取出连接
        @return MySQLdb.connection
        """
        if not DBPool.poolSelectDict.has_key(dbName):
            pool = PooledDB(creator=MySQLdb, mincached=1 , maxcached=1 ,
                              host=DB_HOST , port=DB_PORT , user=DB_USER , passwd=DB_PWD ,
                              db=dbName,use_unicode=False,charset='utf8',cursorclass=DictCursor)
            print 'pool=%s' % pool
            DBPool.poolSelectDict[dbName] = pool
        return DBPool.poolSelectDict[dbName].connection()

# mypool = DBPool.getConn('billing_record_db')
# print 'mypool=%s' % mypool
         
class DB(object):
    """
        MYSQL数据库对象，负责产生数据库连接 , 此类中的连接采用连接池实现
        获取连接对象：conn = Mysql.getConn()
        释放连接对象;conn.close()或del conn
    """
    
    def __init__(self):
        self.pool = DBPool
        self.sql = ""


    def _create(self, dbName):
        '''从连接池拿连接和指针'''
        # con = self.pool.__getConn(dbName)
        con = self.pool.getConn(dbName)
        # con.autocommit(1)
        # cursor = con.cursor( umysqldb.cursors.DictCursor )
        cursor = con.cursor()
        return con,cursor

    def _selectCreate(self, dbName):
        '''从连接池拿连接和指针'''
        # con = self.pool.__getConn(dbName)
        con = self.pool.getSelectConn(dbName)
        # con.autocommit(1)
        # cursor = con.cursor( umysqldb.cursors.DictCursor )
        cursor = con.cursor()
        return con,cursor
    
    def _close(self, con=None, cursor=None):
        '''关闭连接'''
        try:
            if con:
                con.commit()
                # print '--- DB.con > commit   success'

        except Exception,e:
            msg = ' --- DB.con > commit fail ',e
            ERRORLOG.error(msg)
            print msg

        try:
            if cursor:
                cursor.close()
                # print '--- DB.cursor >  close success'
        except Exception,e:
            msg =  ' --- DB.cursor > closed fail ',e
            ERRORLOG.error(msg)
            print msg
        try:
            if con:
                con.close()
                # print '--- DB.con > close success'

        except Exception,e:
            msg = ' --- DB.con > closed fail ',e
            ERRORLOG.error(msg)
            print msg

    def _query(self, dbName, sql, args=None, fetchAll=True, slave=True):
        '''执行查询语句,并返回结果,推荐直接使用fetchone(),fetchall()'''
        try:
            if NODE_NAME == 'erp' and  sql.lower().strip().startswith('select') and slave:
                con,cursor = self._selectCreate(dbName) 
            else:
                con,cursor = self._create(dbName) 

            cursor.execute(sql,args)
            if fetchAll:
                result =  cursor.fetchall()
            else:
                result = cursor.fetchone()

            return result if result else {} #有无记录返回类型一致 ,以免前端闪退
        except Exception, e:
            msg = '%s >>> dbName=%s sql=%s args=%s'%(e,dbName,sql,args)
            ERRORLOG.error(msg)
            print 'utils-DB-_query: msg=%s' % msg
            return False

        finally:
            self.sql = sql
            self._close(con, cursor)

    def _execute(self, dbName,sql,args=None,multiKey=False):
        '''增 删 改
        @note:涉及连接
        @param multiKey: 是否多主键
        @return: 
            -insert返回Id（Insert语句联合主键时候如果没有指定multiKey=True,返回0）
            -其他返回影响行数
        '''
        try:
            # print 'sql=%s' % sql
            con,cursor = self._create(dbName) 
            result = cursor.execute(sql,args) #如果是更新直接就是影响行数

            if sql.lower().strip().startswith('insert') and not multiKey:
                return cursor.lastrowid  #如果是插入语句,且单主键,返回lastid
            else:
                return result
        except Exception, e:
            msg = '%s >>> dbName=%s sql=%s args=%s'%(e,dbName,sql,args)
            ERRORLOG.error(msg)
            print 'utils-DB-_execute: msg=%s' % msg
            return False

        finally:
            self.sql = sql
            self._close(con, cursor)

    def _executemany(self, dbName, sql, args):
        '''原有方法
        @note:涉及连接
        '''
        try:
            con,cursor = self._create(dbName)
            result = cursor.executemany(sql, args)
            
            return result
        except Exception, e:
            msg = '%s >>> dbName=%s sql=%s args=%s'%(e,dbName,sql,args)
            ERRORLOG.error(msg)
            print 'utils-DB-_executemany: msg=%s' % msg
            return False

        finally:
            self.sql = sql
            self._close(con, cursor)

    #================================= 以下方法不涉及连接  ==================================  
    def fetchone(self, dbName, sql, args=None):
        '''查询一条记录  @return dict'''
        result = self._query(dbName,sql,args,fetchAll=False)
        return result if result else {} #有无记录返回类型一样 dict


    def fetchall(self, dbName, sql, args=None):
        '''查询所有记录 @return tuple'''
        result = self._query(dbName,sql,args,fetchAll=True)
        return result if result else () #有无记录返回类型一样 tuple



    def out_field(self, dbName, table, field, where='1'):
        '''field是列名,返回一个值'''
        sql = "SELECT %s FROM %s WHERE %s"%(field,table,where)
        row = self.fetchone(dbName,sql)
        if not row:
            return None
        return row.get(field)

    def count(self, dbName, table, where='1'):
        '''获取数量 @return int'''
        return self.out_field(dbName,table,'count(1)',where)
    
    def out_list(self, dbName, table, field, where='1', distinct=False):
        ''' 查询所有满足条件的多行中的field(有相同或不同) 组成一个列表 
            @return list
            @example:
            out_list('billing_record_db','order_goods','order_no','payment_method=%s',True) #获取订单号
        '''
        dfield = "DISTINCT(%s) AS %s"%(field,field) if distinct else field
        sql = "SELECT %s FROM %s WHERE %s"%(dfield,table,where)
        rows = self.fetchall(dbName,sql)
        rlist = []
        for row in rows:
            rlist.append(row.get(field))
        return rlist

    def out_row(self, dbName, table, fields, where='1'):
        '''
            @param fields是一个列表,包含列名,返回一个字典
            @return dict-从数据库中读出的记录。
        '''
        sfield = ",".join(fields) if isinstance(fields, (list, tuple)) else fields
        sql = "SELECT %s FROM %s WHERE %s"%(sfield,table,where)
        row = self.fetchone(dbName,sql)
        return row

    def out_rows(self, dbName, table, fields, where='1'):
        '''
            @param fields是一个列表,包含列名,
            @return tuple-从数据库中读出的记录。
        '''
        sfield = ",".join(fields) if isinstance(fields, (list, tuple)) else fields
        sql = "SELECT %s FROM %s WHERE %s"%(sfield,table,where)
        # print 'sql=%s' % sql
        rows =  self.fetchall(dbName,sql)
        return rows

    def sort(self,rows,order):
        '''排序
        @param rows: ({},{},)
        @para order: ['-order_no','payment_method',] 按订单号降序,支付方式升序   注意:排序的值降序时必须是整数
        '''
        def makekey(dic):
            t = ()
            for o in order:
                if o.startswith('-'):
                    t+=(dic[o[1:]]*-1,)
                else:
                    t+=(dic[o],)
            return t
        rows = sorted(rows, key = makekey, reverse=False)
        return rows

    def insert(self, dbName, table, row, raw=[],replace=False):
        '''单条插入 @return int or False
            @param row: dict
            @param raw: list 不加引号字段 , 已兼容方法类或字段类的   ins['RecordTime']='Now()' ,raw=['RecordTime'] 
        '''
        fields = []
        values = []
        for k,v in row.iteritems():
            fields.append(k)
            if raw and k in raw or isinstance(v, (int,long,float)):
                values.append(v)
            else:
                v = "'%s'"%self.escape_string(v) if isinstance(v, basestring) else v
                values.append(v)
        
        action = 'REPLACE' if replace else 'INSERT'
        sql = "%s INTO %s (%s) VALUES (%s)"%(action, table, ",".join(fields), ",".join([str(v) for v in values]) ) #值里有%s的有问题
        return self._execute(dbName,sql)

    def update(self, dbName, table, row, where, raw=[]):
        '''执行更新语句  @return int
        @param raw:不加引号字段   已兼容方法类或字段类的   
        @e.g.  ups['cargo_fee_rate']='cargo_fee_rate+0.1',raw=['cargo_fee_rate',] 
        '''
        groups = []
        for k,v in row.iteritems():
            if k in raw:
                groups.append("%s=%s"%(k,v))
            else:
                cell = "%s='%s'"%(k,self.escape_string(v)) if isinstance(v, basestring) else "%s=%s"%(k,v)
                groups.append(cell)
        sql = "UPDATE %s SET %s WHERE %s"%(table, ",".join(groups), where)
        return self._execute(dbName,sql)
        
    def insert_update(self, dbName, table, arr, arr_check):
        '''存在则更新，不存在则插入'''
        where = self.mkWhere(arr_check)
        stat = self.count(dbName,table, where)
        if stat:
            update_arr = dict(list(set(arr.items()) - set(arr_check.items())))
            return self.update(dbName,table,update_arr,where) 
        return self.insert(dbName,table,arr)

    def mkWhere(self,wdict):
        '''由字典组成 where语句 {'a':1,'b':'c'} 转成  a=1 AND b='c'  '''
        where = ''
        groups = []
        for k,v in wdict.iteritems():
            cell = "%s='%s'"%(k,self.escape_string(v)) if isinstance(v, basestring) else "%s=%s"%(k,v)
            groups.append(cell)
            where = ' AND '.join(groups)
        return where
    
    def insertmany(self, dbName, table, rows):
        '''批量插入 rows: [{'id':1,'val':1},{'id':2,'val':2},..] 对应于out_rows() '''
        fields = rows[0].keys()
        values = [row.values() for row in rows] #暂不做值的escape_string
        vlist = ',%s'*len(fields)
        sql = 'INSERT INTO %s (%s) VALUES(%s)'%(table, ','.join(fields), vlist[1:])
        return self._executemany(dbName,sql,values)
        
        
    def insertmanynew(self, dbName, table, rows, raw=[]):
        '''批量插入 rows: [{'id':1,'val':1},{'id':2,'val':2},..] 对应于out_rows() '''
        
        for row in rows:
            for k,v in row.iteritems():
                if isinstance(v, (datetime.datetime,datetime.date,datetime.time)):
                    v = v.strftime('%Y-%m-%d %H:%M:%S')
                    v = "'%s'"%v
                elif raw and k in raw or isinstance(v, (int,long,float)):
                    continue
                elif isinstance(v, basestring):
                    v = "'%s'"%self.escape_string(v)
                elif v is None:
                    v = "''"
                
                row[k] = v

        result = 0 #结果
        tLength = len(rows) #个数
        tMax = 5000 #最大个数 
        if tLength>tMax:
            num = tLength / tMax
            for index in range(num+1):
                newRows = rows[index*tMax:(index+1)*tMax]
                if newRows:
                    sql = self.getInsertManySql(table, newRows)
                    result += (self._execute(dbName,sql,multiKey=True) or 0)
        else:
            sql = self.getInsertManySql(table, rows)
            result = self._execute(dbName,sql,multiKey=True)
        return result
    
    def getInsertManySql(self,table,rows):
        '''获取insertmany sql语气'''
        fields = rows[0].keys()
        sql = 'INSERT INTO %s (%s) VALUES '%(table, ','.join(fields))
        values = []
        for row in rows:
            value = "(" + ",".join([str(v) for v in row.values()]) + ")"
            values.append(value)
        sql += ",".join(values)
        return sql

    def updatemany(self, dbName, table, info, where):
        '''一次更新多条记录 谨用 
        @param table: 表名
        @param where: 条件(可能跟要更新的内容有关，暂自行构造 )
        @example: 
        #单字段更新 (常用)
        info = ('field1','field2',{when1:then1,when2:then2,..},'field1') //最后元素缺省为第一个元素
        #说明:Info会组成  SET 'field1' = CASE 'field2' When k Then v.. ELSE 'field1' END 
        #多单段更新 (少见)
        info = [('field1','field2',{1:2,2:3,3:4,4:5}), ('val3','val4',{1:0,2:1,3:2},'val1')] 
        where = 'id in (1,2,3,4)'
        @note: 如果where中条件包含 又没有给出WHEN THEN的值 就会改为ELSE的值,如果没有给出ELSE的值便是0或空
        @remark: 已使用UPDATE good SET price = CASE goodId WHEN 1001 THEN price+57018 WHEN 1002..
        '''
        assert isinstance(info, (list,tuple))
        
        if isinstance(info,tuple):
            info = [info]
        sql = 'UPDATE %s SET'%table
        for row in info: 
            sql += ' %s = CASE %s' % (row[0], row[1])
            for k,v in row[2].iteritems():
                sql += ' WHEN %s THEN %s' % (k,v)
            elsevalue = row[3] if len(row)>3 else row[0] #暂只支持字段或整数- 不支持字符串 因为无引号
            sql += ' ELSE %s' % elsevalue
            sql += ' END,' #如果where中有包含 但又没给出when then则会变为0, 如果这种情况用ELSE 原字段 END
        sql = sql.strip(',') + ' WHERE %s'%where
        return self._execute(dbName,sql)

    def delete(self, dbName, table,where):
        '''删除记录 '''
        sql = "DELETE FROM %s WHERE %s"%(table,where)
        return self._execute(dbName,sql)

    def inWhere(self,field,inList,notIn=False,symbol=False):
        '''
            @param field: 字段
            @param inList: 列表 
            @param symbol: 是否有引号
            @return string 条件字符串,用于构造条件
            @example:
            ('id',[1,2,3])  >> id in ('1','2','3') or id in (1,2,3)
            ('Name',['a','b','c'])  >> Name in ('a','b','c')
        '''
        assert isinstance(field, str)
        assert isinstance(inList, (list,tuple))
        if not inList: #对空列表特殊处理
            inList = ['@']
            symbol=True
        inList = [str(cell) for cell in inList]
        notSign = ' NOT' if notIn else ''
        if symbol:
            where = field+notSign+" IN ('"+str.join("','",inList)+"')"
        else:
            where = field+notSign+" IN ("+str.join(",",inList)+")"

        return where

    def escape_string(self,s):
        '''转义逃脱sql中的特殊字符'''
        return MySQLdb.escape_string(s) 

ObjDB = DB() #db对象
# print 'objDB=%s' % ObjDB

class MyDefineError(Exception):
    pass

class MyDefineErrorWithStatusCode(Exception):
    pass

from tornado.web import RequestHandler


class BasicTemplateHandler(RequestHandler):

    version = VERSION

    WAP_HOST = WAP_HOST

    WEB_HOST = WEB_HOST

    # client = CLIENT
    client = getClient() 

    expires_days = WEB_COOKIE_EXPIRE_DAYS if CLIENT == "web" else WAP_COOKIE_EXPIRE_DAYS

    record_operate_log = True

    def __init__(self, *request, **kwargs):
        
        # if self.client == 'wap':
            # self.expires_days = WAP_COOKIE_EXPIRE_DAYS
        # else:
            # self.expires_days = WEB_COOKIE_EXPIRE_DAYS
        super(BasicTemplateHandler, self).__init__(*request, **kwargs)
        
        # print self.request.path
        if self.request.path == "/" or self.request.uri.startswith("/show/") \
            or self.request.uri.startswith("/detail/"):
            r = getRedisObj()
            if self.request.path == "/":
                last_visit = self.get_secure_cookie("last_visit")
                today = str(datetime.date.today())
                if last_visit != today:
                    self.set_secure_cookie("last_visit",today,expires_days=10)
                    source = self.get_argument("from","") or self.client
                    if source == "xiaoherios":
                        source = "ios"
                    elif source == "xiaoherandroid":
                        source = "android"
                    else:
                        source = self.client
                    key = "%s:%s"%(today,source)
                    # print key,222222222222222
                    r.incr(key)

            uid = self.get_current_user()
            score = int(time.time()) + 10 * 60
            r.zadd("ten_online", uid, score)


    def compute_etag(self):
        return None

    def get_argument(self, name, default=[], strip=True):
        _argument = RequestHandler.get_argument(self, name, default, strip=strip)

        #The returned value of RequestHandler.get_argument is always unicode if no default given.
        #But if default is given,return could be anything.
        #Only filter unicode here.
        if 1 == 2 and isinstance(_argument, type(u"1")):
            #replac "1 == 1" with "switch" here
            if arguementFiltered == "":
                arguementFiltered = default

            if arguementFiltered != _argument:
                logging.warning("Sql injection deteceted. Attack url:[{0}] Data:[ {1} : {2} ]".format(self.request.uri,
                                                                                                      name,
                                                                                                      _argument))
            return arguementFiltered
        return _argument


    def getArgument(self):

        kws = {}
        rest_argument = self.restfulArgument()

        if not rest_argument:
            return kws

        for i in rest_argument:
            
            length = len(i)
            if length == 4:
                ##默认值
                value = self.get_argument(i[0],i[3])
            else:
                value = self.get_argument(i[0])
            
            if length>=2 and i[1]:
                ###初始化函数
                value = i[1](value)

            if length >= 3 and i[2]:
                ###重命名参数
                kws[i[2]] = value
            else:
                kws[i[0]] = value

        return kws

    def restfulArgument(self):
        '''
        [("name",None),("age",int)]
        '''
        return []

    def getAllArguments(self):
        argus = self.request.arguments
        for key in argus:
            argus[key] = self.get_argument(key)
        return argus

    def DBAction(self, Arguments):
        # remember to self.write to client
        pass

    def DBAction_post(self, Arguments):
        # remember to self.write to client
        pass


    def countAPIclickTime(self):
        source = self.get_secure_cookie("source")
        if source:
            pass
        else:
            source = self.client
        if self.request.uri.startswith("/show/"):
            api_path = "/show/"
            full_path = self.request.path
        elif self.request.uri.startswith("/detail/"):
            api_path = "/detail/"
            full_path = self.request.path
        elif self.request.uri.startswith("/fuli/"):
            api_path = "/fuli/"
            full_path = self.request.path
        else:
            api_path = self.request.path
            full_path = self.request.path
        date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        uid = self.get_current_user()
        if uid:
            value = {"date":date,"uid":uid,"source":source,"api_path":api_path,"full_path":full_path}
            r = getRedisObj(15)
            r.lpush("api_list",value)

    def EscapeSQL(self):
        return True

    def _post(self, **kws):
        try:
            Arguments = self.getArgument()
            assert isinstance(Arguments, dict) == True
            if self.EscapeSQL():
                for key, value in Arguments.items():
                    try:
                        value = MySQLdb.escape_string(value)
                        Arguments[key] = value
                    except:
                        pass
            Arguments.update(kws)
            try:
                self.DBAction(Arguments)
                if self.record_operate_log:
                    # logStr = self.get_operate_log(Arguments)
                    # logStr and DEBUGLOG.debug(logStr)
                    # print logStr

                    # --------------------------------Modify bu cyf 2015-01-14 write into DB-------------
                    operator, operUrl, operArgu, operModule = self.get_operate_log(Arguments)

                    if operator:
                        now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        db = DBAccess()
                        db.dbName = "log_records"

                        sql = " insert into operation_records (operator,operate_time,operate_module,url,arguments) values('%s','%s','%s','%s','%s')"%(operator,now_time,operModule,operUrl,operArgu)
                        try:
                            db.execNonQuery(sql)
                        except MySQLdb.Error, e:
                            pass

            except MyDefineError as e:
                self.write({"status": RET_DBERROR, "msg": str(e)})
                # errorStr = traceback.format_exc()
                # ERRORLOG.error(errorStr)

            except MyDefineErrorWithStatusCode as e:
                self.write({"status": e[0], "msg": e[1]})
                # errorStr = traceback.format_exc()
                # ERRORLOG.error(errorStr)  

            except ValueError:
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)
                self.write("参数不对")
                              
            except Exception as e:
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)
                if MAIL_WARN: #and self.request.uri.startswith("/orders/add"):
                    send_mail(self.request.uri, errorStr, ADMINS )
                self.write({"status": RET_DBERROR, "msg":"出错拉"})
        except:
            errorStr = traceback.format_exc()
            ERRORLOG.error(errorStr)
            self.write({"status": RET_HTTP_PARAMETER_ERROR})

    def _post_pages(self, **kws):
        try:
            Arguments = self.getArgument()
            assert isinstance(Arguments, dict) == True
            if self.EscapeSQL():
                for key, value in Arguments.items():
                    try:
                        value = MySQLdb.escape_string(value)
                        Arguments[key] = value
                    except:
                        pass
            Arguments.update(kws)
            try:
                self.DBAction_post(Arguments)
                if self.record_operate_log:
                    # logStr = self.get_operate_log(Arguments)
                    # logStr and DEBUGLOG.debug(logStr)
                    # print logStr

                    # --------------------------------Modify bu cyf 2015-01-14 write into DB-------------
                    operator, operUrl, operArgu, operModule = self.get_operate_log(Arguments)

                    if operator:
                        now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        db = DBAccess()
                        db.dbName = "log_records"

                        sql = " insert into operation_records (operator,operate_time,operate_module,url,arguments) values('%s','%s','%s','%s','%s')"%(operator,now_time,operModule,operUrl,operArgu)
                        try:
                            db.execNonQuery(sql)
                        except MySQLdb.Error, e:
                            pass

            except MyDefineError as e:
                self.write({"status": RET_DBERROR, "msg": str(e)})
                # errorStr = traceback.format_exc()
                # ERRORLOG.error(errorStr)

            except MyDefineErrorWithStatusCode as e:
                self.write({"status": e[0], "msg": e[1]})
                # errorStr = traceback.format_exc()
                # ERRORLOG.error(errorStr)  

            except ValueError:
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)
                self.write("参数不对")
                              
            except Exception as e:
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)
                self.write({"status": RET_DBERROR})
        except:
            errorStr = traceback.format_exc()
            ERRORLOG.error(errorStr)
            self.write({"status": RET_HTTP_PARAMETER_ERROR})

    
    def get_human_param(self,Arguments):
        logStr = ""
        for key,value in Arguments.items():
            logStr += "%s=%s&"%(key,value)
        return logStr

    def get_operate_log(self,Arguments):
        url = self.request.uri 
        if url.startswith("/admin/orders/change"):
            from shop_admin.models import insertOrderOperateRecord
            from orders.models import record_order_flow
            order_id = Arguments['order_id']
            if isinstance(order_id,int):
                order_id = [order_id]
            status = Arguments["status"]
            operator = self.get_secure_cookie("cuser")
            [ insertOrderOperateRecord(order_id=i,order_goods_id=0,set_status=status,operator=operator) for i in order_id]
            [ record_order_flow(order_no=i,order_goods_id=0,operator=operator,type=status,remark="admin_"+status) for i in order_id]

        elif url.startswith("/admin/order/goods/change"):
            from shop_admin.models import insertOrderOperateRecord
            from orders.models import record_order_flow
            order_goods_id = Arguments['order_goods_id']
            status = Arguments["status"]
            operator = self.get_secure_cookie("cuser")
            [ insertOrderOperateRecord(order_id=0,order_goods_id=i,set_status=status,operator=operator) for i in order_goods_id]

        else:
            pass

        if url.startswith("/company") \
            or url.startswith("/admin"):
            company_id = self.get_secure_cookie("cuser")
            humanParam = self.get_human_param(Arguments)
            #logStr = "%s %s %s "%(company_id,self.request.uri,humanParam)
            #return logStr

            operUrl = ""
            if self.request.uri[-1] == '/':
                operUrl = self.request.uri[:len(self.request.uri)-1]
            else:
                operUrl = self.request.uri
            belong_module = ""
            group =  list(getMongoDBConn().shop.permission.find({"url": operUrl}, {"tag": 1}))

            if len(group) > 0:
                belong_module = group[0].get("tag","未分类")
            else:
                belong_module = "未分类"

            return company_id, self.request.uri, humanParam, belong_module

        elif  self.request.uri.startswith("/orders/add") :
            uid = self.get_current_user()
            humanParam = self.get_human_param(Arguments)
            #logStr = "%s %s %s "%(uid,self.request.uri,humanParam)
            #return logStr

            operUrl = ""
            if self.request.uri[-1] == '/':
                operUrl = self.request.uri[:len(self.request.uri)-1]
            else:
                operUrl = self.request.uri
            belong_module = ""
            group =  list(getMongoDBConn().shop.permission.find({"url": operUrl}, {"tag": 1}))
            if len(group) > 0:
                belong_module = group[0].get("tag","未分类")
            else:
                belong_module = "未分类"

            return uid, self.request.uri, humanParam, belong_module

        return "", "", "", ""

    def _get(self, **kws):
        return self._post(**kws)

    def get_current_user(self):
        user = self.get_secure_cookie("user")
        try:
            return int(user)
        except:
            return 0

    def is_virtual_user(self):
        user_id = self.get_current_user()
        if user_id == 0:
            return True
        return user_id >= MIN_VIRTUAL_USER_ID

    def get_user_porm(self):
        porm = self.get_secure_cookie("porm")
        return porm or ""

    def get_company_info(self):
        company_name = self.get_secure_cookie("company_name")
        level = self.get_secure_cookie("level")
        try:
            level = int(float(level))
        except:
            level = -1
        # print level
        return {"company_name": company_name, "level": level}

    def get_define_size_image(self, url, mode=2, w=0, h=0):
        return getDefineSizeImage(url, mode, w, h)

    def get_client_type(self):
        try:
            ua = self.request.headers["User-Agent"].lower()
        except:
            return "web"
        if PHONE_UA_RE.search(ua):
            return "wap"
        else:
            return 'web'

    def write_error(self, status_code, **kwargs):
        if self.request.method not in("GET","POST"):
            self.redirect("http://www.gov.cn/")
            return
        if status_code == 405:
            self.redirect("http://www.xiaoher.com")
            return
        elif status_code == 400:
            self.write("参数错误")
            return
        import traceback
        # if self.settings.get("debug") and "exc_info" in kwargs:
        html = ""
        if "exc_info" in kwargs:
            exc_info = kwargs["exc_info"]
            trace_info = ''.join(
                ["%s<br/>" % line for line in traceback.format_exception(*exc_info)])
            request_info = ''.join(["<strong>%s</strong>: %s<br/>" % (k, self.request.__dict__[k])
                                   for k in self.request.__dict__.keys()])
            error = exc_info[1]

            html = """<html><title>%s</title>
                    <body>
                        <h2>Error</h2>
                        <p>%s</p>
                        <h2>Traceback</h2>
                        <p>%s</p>
                        <h2>Request Info</h2>
                        <p>%s</p>
                        <h2>user</h2>
                        <p>%s</p>
                        </body></html>""" % (error, error, trace_info, request_info,self.get_current_user())

        if self.settings.get("debug"):
            self.set_header('Content-Type', 'text/html')
            self.finish(html)
        else:
            # to do : mail warning
            if html and MAIL_WARN:
                send_mail(self.request.uri, html, ADMINS, mail_type="html")

            # for line in traceback.format_exception(*exc_info):
            ERRORLOG.error('\n'.join(traceback.format_exception(*exc_info)))
            
            self.set_status(status_code)
            self.render("error/error.html")

    def on_finish(self):

        finish_time = self.request.request_time()
        if finish_time > 2:
            db = getMongoDBConn().request_warn
            db.request_warn.save({"url":self.request.uri,"finish_time":finish_time,"date_time":getNowUTCtime()})



class AsyncHandler(BasicTemplateHandler):

    @tornado.gen.coroutine
    def _post(self, **kws):
        try:
            Arguments = self.getArgument()
            assert isinstance(Arguments, dict) == True
            if self.EscapeSQL():
                for key, value in Arguments.items():
                    try:
                        value = MySQLdb.escape_string(value)
                        Arguments[key] = value
                    except:
                        pass
            Arguments.update(kws)
            try:
                yield self.DBAction(Arguments)
                if self.record_operate_log:
                    # logStr = self.get_operate_log(Arguments)
                    # logStr and DEBUGLOG.debug(logStr)
                    # print logStr

                    # --------------------------------Modify bu cyf 2015-01-14 write into DB-------------
                    operator, operUrl, operArgu, operModule = self.get_operate_log(Arguments)

                    if operator:
                        now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        db = DBAccess()
                        db.dbName = "log_records"

                        sql = " insert into operation_records (operator,operate_time,operate_module,url,arguments) values('%s','%s','%s','%s','%s')"%(operator,now_time,operModule,operUrl,operArgu)
                        try:
                            db.execNonQuery(sql)
                        except MySQLdb.Error, e:
                            pass

            except MyDefineError as e:
                self.write({"status": RET_DBERROR, "msg": str(e)})
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)

            except MyDefineErrorWithStatusCode as e:
                self.write({"status": e[0], "msg": e[1]})

            except ValueError:
                self.write("参数不对")

            except Exception as e:
                errorStr = traceback.format_exc()
                ERRORLOG.error(errorStr)
                self.write({"status": RET_DBERROR})
        except:
            errorStr = traceback.format_exc()
            ERRORLOG.error(errorStr)
            self.write({"status": RET_HTTP_PARAMETER_ERROR})       


def dictFilter(_dict):
    '''字典过滤'''
    assert isinstance(_dict, dict) == True
    d = copy.deepcopy(_dict)
    for k in d.keys():
        if not d[k]:
            d.pop(k)
    return d


def getNowUTCtime(utc=False):
    if utc:
        nowTime = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    else:
        nowTime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return nowTime

def getNowUTCDate(utc=False):
    if utc:
        nowDate = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    else:
        nowDate = datetime.datetime.now().strftime("%Y-%m-%d")
    return nowDate


def getCodeMsg(result):
    status = result["status"]
    if status != 0:
        try:
            result["msg"] = MSG_CODE[status]
        except:
            pass
    return result


class MessagesCenter:

    #temporary solution.
    @staticmethod
    def initRegister():
        MessagesCenter.register("ProductReturned","shop_admin.models.handleIncomingStockMessage")
        MessagesCenter.register("ProductDeclined","shop_admin.models.handleIncomingStockMessage")
        MessagesCenter.register("PurchaseFinished","shop_admin.models.handleIncomingStockMessage")
        MessagesCenter.register("PurchaseFailed","shop_admin.models.handleIncomingStockMessage")
        MessagesCenter.register("OrderFinished","orders.models.handleOrderFinishedMessage")

    @staticmethod
    def sendMessage(messageName,paras):
        MessagesCenter.initRegister()

        r = getRedisObj(rdb = 3)
        funcs = r.smembers("MESSAGE_%s" % messageName)


        for item in funcs:
            print item
            parts = item.split(".")
            if len(parts) < 3:
                ERRORLOG.error("ERROR IN sendMessage:%s" % messageName)
                return 
            importStr = "from %s.%s import %s" % (parts[0],parts[1],parts[2],)
            exec(importStr)
            func = eval(parts[2])
            func(paras)

    @staticmethod
    def register(messageName,funcName):
        r = getRedisObj(rdb = 3)
        r.sadd("MESSAGE_%s" % messageName , funcName)
        pass

    @staticmethod
    def deregister(messageName,funcName):
        r = getRedisObj(rdb = 3)
        r.srem("MESSAGE_%s" % messageName , funcName)


def incrForkey(key):
    conn = getMongoDBConn()
    db = conn.shop
    query = {"_id":key}
    update = {"$inc":{"seq":1}}
    obj = db.counters.find_and_modify(query,update,True,new=True)
    return obj["seq"]



def paginationStartEnd(page_index, page_max):
    if page_index <= 6:
        return 1, min(page_max, 10)

    if page_max <= 10:
        page_start = 1
        page_end = page_max
        return page_start, page_end

    page_end = min(page_max, page_index+5)

    page_start = page_end - 9
    return page_start, page_end

def open_excel(file=''):
    try:
     data = xlrd.open_workbook(file)
     return data
    except Exception,e:
     print str(e)

def excel_table_byindex(file='',colnameindex=0,index=0):

    data = xlrd.open_workbook(file_contents=file['body'],encoding_override='windows-1252')
    table = data.sheet_by_index(index)
    nrows = table.nrows #行数
    ncols = table.ncols #列数
    colnames =  table.row_values(colnameindex) #某一行数据
    list =[]
    for rownum in range(1,nrows):
      row = table.row_values(rownum)
      list.append(row)

    return list

def excel_table_byindex_xlsx(filename, sheetnames=[], row_index=1, col_index=1):
    '''读取2007 以上的xlsx 文件
        @param filename:包含路径的文件名
        @param sheet_name:要读取的表名数组
        @param row_index:从第几行开始
        @param col_index:从第几列开始
        @return dict : {表名:[[cell,]] 表数据}
    '''
    wb = load_workbook(filename = filename )  
    if not sheetnames:
        sheetnames = wb.get_sheet_names()  
    tDict = {}
    for sheetname in sheetnames:
        ws = wb.get_sheet_by_name(sheetname)   
        rList = []
        rowNum = ws.get_highest_row()
        colNum = ws.get_highest_column()  
        for rn in xrange(row_index,rowNum):
            cList = []
            for cn in xrange(col_index,colNum):
                cellValue = ws.cell(row=rn,column=cn).value
                if cn!=col_index and not cellValue:
                    break
                cList.append(cellValue)
                # print cellValue
            rList.append(cList)
        tDict[sheetname] = rList
    return tDict

def getRandStr(n):
    '''获取随机字符串'''
    st = ''
    randStr = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    for i in range(n):
        st += random.choice(randStr)
    return st



if __name__ == '__main__':
    dbName = 'billing_record_db'
    table = 'stock'
    print  ObjDB.out_row(dbName,'address','*')
    # print ObjDB.out_row(dbName,'stock','*')
    # row = ObjDB.out_row(dbName,'stock',['id','uid'])
    # print 'row=%s' % row
    # where = ObjDB.inWhere('id',[11,12,13])
    # print ObjDB.out_rows(dbName,'stock',['id','uid'],where)
    # print ObjDB.out_rows(dbName,'stock',['id','uid'],where,order=['-id'])
    # print 'rows=%s' % list(rows)
    # print ObjDB.out_field(dbName,'stock','uid',where)
    # print ObjDB.out_field(dbName,'stock','id',where)
    # print ObjDB.out_list(dbName,'stock','uid',where)
    # print ObjDB.count(dbName,'stock')
    # import datetime 
    # stockDict = {'purchase_id': 0L, 
    # 'remark': '\xe5\x8f\x91\xe5\xb8\x83\xe5\x88\xb0\xe6\x9f\x90\xe6\xb4\xbb\xe5\x8a\xa8\xe4\xbd\xa0\xe5\xa6\xb9\xe5\x95\x8a',
    #  'order_goods_id': 176L, 'uid': 7L, 
    #  'order_no': '0', 
    #  'destination': '\xe5\x8f\x91\xe5\xb8\x83\xe5\x88\xb0\xe6\x9f\x90\xe6\xb4\xbb\xe5\x8a\xa8', 
    #  'update_time':'2014-05-30 18:22:51', 
    #  'source': '\xe9\x87\x87\xe8\xb4\xad\xe5\xae\x8c\xe6\x88\x90', 
    #  'create_time': '2014-05-30 14:39:25', 'depot': 'C702'}

    # stockDict2 = {'purchase_id': 0L, 
    # 'remark': '\xe5\x8f\x91\xe5\xb8\x83\xe5\x88\xb0\xe6\x9f\x90\xe6\xb4\xbb\xe5\x8a\xa8\xe4\xbd\xa0\xe5\xa6\xb9\xe5\x95\x8a',
    #  'order_goods_id': 176L, 'uid': 7L, 
    #  'order_no': '0', 
    #  'destination': '\xe5\x8f\x91\xe5\xb8\x83\xe5\x88\xb0\xe6\x9f\x90\xe6\xb4\xbb\xe5\x8a\xa8', 
    #  'update_time':'2014-05-30 18:22:51', 
    #  'source': '\xe9\x87\x87\xe8\xb4\xad\xe5\xae\x8c\xe6\x88\x90', 
    #  'create_time': '2014-05-30 14:39:25', 'depot': 'C702'}
    # print ObjDB.insert(dbName,'stock',stockDict)
    # print ObjDB.update(dbName,'stock',{'uid':'uid+1'},'id=86',['uid'])
    # print ObjDB.update(dbName,'stock',{'uid':5},'id=86')
    # print ObjDB.insert_update(dbName, table, stockDict, arr_check={'id':90})
    # import copy
    # print ObjDB.insertmany(dbName,table,[stockDict,stockDict2])
    # print ObjDB.insertmanynew(dbName,table,[stockDict,stockDict2])
    # info = [('uid','id',{99:1,98:2,97:3})]
    # where = ObjDB.inWhere('id',[99,98,97])
    # print ObjDB.updatemany(dbName,table,info,where)
    # where = 'id > 85'
    # print ObjDB.delete(dbName,table,where)
    # f = {'body':'/home/weigenming/Downloads/顺丰派送范围1.xls'}
    # f = {'body':'/home/weigenming/Downloads/xiao4.xlsx'}
    f = '/home/weigenming/Downloads/快递.xlsx'
    # print excel_table_byindex(f)
    # tDict = excel_table_byindex_xlsx(f,['顺风陆运'],2,2)
    # print tDict.values()[0][0]
    # print tDict.values()[0][1]
    # print tDict.values()[0][2]
    # xiao = u'\u5b89\u5e86\u5e02'
    # print xiao
    gen = u'\u868c\u57e0\u5e02'
    print gen
    print getRandStr(5)
    # print 'tDict=%s' % tDict

    




    

